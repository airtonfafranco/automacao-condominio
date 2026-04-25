import fitz
from PIL import Image, ImageDraw, ImageFont
import os
import unicodedata
from pathlib import Path
import pandas as pd
import re
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


# ==============================================================
# CONSTANTES DE NEGÓCIO — não mudam entre execuções
# ==============================================================
SENHAS_PDFS = {
    "claro internet": "00978",
    "claro chip": "00978",
    "light": "0097",
}

REGRA_PADRAO_CONTA = {
    "pagina": 0,
    "y0": 0.00,
    "y1": 1.00,
    "redimensionar": 0.50,
}

REGRAS_CONTAS_ESPECIAIS = {
    "claro internet": {
        "pagina": 1,
        "redimensionar": 0.50,
    },
}

REGRA_COMPROVANTE = {
    "pagina": 0,
    "y0": 0.00,
    "y1": 0.60,
}


# ==============================================================
# FUNÇÕES — ETAPA 1: AGRUPAMENTO
# ==============================================================

def normalizar_texto(texto):
    texto = texto.lower().strip()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return texto


def obter_nome_base_conta(nome_arquivo):
    nome = Path(nome_arquivo).stem
    if nome.lower().endswith("_conta"):
        nome = nome[:-6]
    return normalizar_texto(nome)


def obter_nome_base_comprovante(nome_arquivo):
    nome = Path(nome_arquivo).stem
    if nome.lower().endswith("_comprovante"):
        nome = nome[:-12]
    return normalizar_texto(nome)


def pdf_para_imagem(caminho_pdf, nome_base, pagina_idx=0, y0=0.0, y1=1.0):
    print(f"   Abrindo PDF: {caminho_pdf}")

    try:
        doc = fitz.open(caminho_pdf)
    except Exception as e:
        raise RuntimeError(f"Erro ao abrir PDF: {caminho_pdf} | Detalhe: {e}")

    try:
        if doc.is_closed:
            raise RuntimeError(f"O documento foi aberto, mas está fechado: {caminho_pdf}")

        if doc.needs_pass:
            senha = SENHAS_PDFS.get(nome_base)
            if not senha:
                raise RuntimeError(f"Senha não cadastrada para o PDF protegido: {caminho_pdf}")
            ok = doc.authenticate(senha)
            if not ok:
                raise RuntimeError(f"Senha incorreta para o PDF protegido: {caminho_pdf}")
            print(f"   PDF desbloqueado com sucesso: {caminho_pdf}")

        if len(doc) == 0:
            raise RuntimeError(f"O PDF não possui páginas: {caminho_pdf}")

        if pagina_idx >= len(doc):
            raise RuntimeError(
                f"O PDF {caminho_pdf} não possui a página solicitada. "
                f"Páginas disponíveis: {len(doc)} | Página pedida: {pagina_idx + 1}"
            )

        pagina = doc.load_page(pagina_idx)
        rect = pagina.rect

        clip = fitz.Rect(
            rect.x0,
            rect.y0 + (rect.height * y0),
            rect.x1,
            rect.y0 + (rect.height * y1),
        )

        mat = fitz.Matrix(5, 5)
        pix = pagina.get_pixmap(matrix=mat, clip=clip)
        return Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

    finally:
        doc.close()


def redimensionar_imagem(img, escala):
    if escala is None or escala == 1:
        return img
    nova_largura = int(img.width * escala)
    nova_altura = int(img.height * escala)
    return img.resize((nova_largura, nova_altura), Image.Resampling.LANCZOS)


def unir_imagens(img_topo, img_base, proporcao_topo=0.70, proporcao_base=0.30):
    largura_final = max(img_topo.width, img_base.width)
    altura_final = img_topo.height + img_base.height

    altura_topo_area = int(altura_final * proporcao_topo)
    altura_base_area = altura_final - altura_topo_area

    nova_img = Image.new("RGB", (largura_final, altura_final), (255, 255, 255))

    def ajustar_para_area(img, largura_max, altura_max):
        escala = min(largura_max / img.width, altura_max / img.height)
        return img.resize(
            (int(img.width * escala), int(img.height * escala)),
            Image.Resampling.LANCZOS,
        )

    img_topo_ajustada = ajustar_para_area(img_topo, largura_final, altura_topo_area)
    img_base_ajustada = ajustar_para_area(img_base, largura_final, altura_base_area)

    x_topo = (largura_final - img_topo_ajustada.width) // 2
    x_base = (largura_final - img_base_ajustada.width) // 2
    y_topo = (altura_topo_area - img_topo_ajustada.height) // 2
    y_base = altura_topo_area + ((altura_base_area - img_base_ajustada.height) // 2)

    nova_img.paste(img_topo_ajustada, (x_topo, y_topo))
    nova_img.paste(img_base_ajustada, (x_base, y_base))

    return nova_img


def extrair_texto_pdf(caminho_pdf, nome_base):
    try:
        doc = fitz.open(caminho_pdf)

        if doc.needs_pass:
            senha = SENHAS_PDFS.get(nome_base)
            if senha:
                if not doc.authenticate(senha):
                    doc.close()
                    return ""
            else:
                doc.close()
                return ""

        textos = [doc.load_page(i).get_text("text") for i in range(len(doc))]
        doc.close()
        return "\n".join(textos)

    except Exception:
        return ""


def extrair_datas_do_texto(texto):
    return re.findall(r"\b(\d{2}/\d{2}/\d{4})\b", texto)


def extrair_valores_do_texto(texto):
    return re.findall(r"\b\d{1,3}(?:\.\d{3})*,\d{2}\b", texto)


def valor_brasileiro_para_float(valor_str):
    try:
        return float(valor_str.replace(".", "").replace(",", "."))
    except Exception:
        return None


def extrair_dados_conta_sem_comprovante(texto_conta):
    texto_limpo = re.sub(r"\s+", " ", texto_conta or "")

    valor = None
    data_pagamento = None

    m_valor = re.search(r"TOTAL\s+A\s+PAGAR\s*\(R\$\)", texto_limpo, flags=re.IGNORECASE)
    if m_valor:
        trecho = texto_limpo[m_valor.end():m_valor.end() + 80]
        valores = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}", trecho)
        if valores:
            valor = valor_brasileiro_para_float(valores[0])

    m_data = re.search(r"VENCIMENTO\s+TOTAL\s*\(R\$\)", texto_limpo, flags=re.IGNORECASE)
    if m_data:
        trecho = texto_limpo[m_data.end():m_data.end() + 80]
        datas = re.findall(r"\d{2}/\d{2}/\d{4}", trecho)
        if datas:
            data_pagamento = datas[0]

    if valor is not None and data_pagamento is not None:
        status = "ok"
    elif valor is not None or data_pagamento is not None:
        status = "parcial"
    else:
        status = "pendente"

    return {"valor": valor, "data_pagamento": data_pagamento, "status_extracao": status}


def salvar_texto_debug(nome_arquivo, conteudo, pasta_debug: Path):
    pasta_debug.mkdir(parents=True, exist_ok=True)
    caminho_saida = pasta_debug / nome_arquivo
    caminho_saida.write_text(conteudo or "", encoding="utf-8")
    print(f"   Debug salvo em: {caminho_saida}")


def extrair_dados_despesa(caminho_conta, caminho_comprovante, nome_base, pasta_debug: Path):
    texto_conta = extrair_texto_pdf(caminho_conta, nome_base)

    if not caminho_comprovante:
        if nome_base == "aguas do rio":
            salvar_texto_debug("aguas_do_rio_conta.txt", texto_conta, pasta_debug)

        dados = extrair_dados_conta_sem_comprovante(texto_conta)
        print(f"   Extração sem comprovante → valor={dados['valor']} | "
              f"data={dados['data_pagamento']} | status={dados['status_extracao']}")
        return dados

    texto_comprovante = extrair_texto_pdf(caminho_comprovante, nome_base)

    if nome_base.lower() == "prolabore":
        salvar_texto_debug("prolabore_conta.txt", texto_conta, pasta_debug)
        salvar_texto_debug("prolabore_comprovante.txt", texto_comprovante, pasta_debug)

    datas_comprovante = extrair_datas_do_texto(texto_comprovante)
    valores_comprovante = extrair_valores_do_texto(texto_comprovante)
    datas_conta = extrair_datas_do_texto(texto_conta)
    valores_conta = extrair_valores_do_texto(texto_conta)

    valor = valor_brasileiro_para_float(valores_comprovante[0]) if valores_comprovante else None
    data_pagamento = datas_comprovante[0] if datas_comprovante else None

    if valor is None and valores_conta:
        valor = valor_brasileiro_para_float(valores_conta[0])
    if data_pagamento is None and datas_conta:
        data_pagamento = datas_conta[0]

    if valor is not None and data_pagamento is not None:
        status = "ok"
    elif valor is not None or data_pagamento is not None:
        status = "parcial"
    else:
        status = "pendente"

    return {"valor": valor, "data_pagamento": data_pagamento, "status_extracao": status}


def obter_regra_conta(nome_base):
    regra = REGRA_PADRAO_CONTA.copy()
    if nome_base in REGRAS_CONTAS_ESPECIAIS:
        regra.update(REGRAS_CONTAS_ESPECIAIS[nome_base])
    return regra


def criar_imagem_aviso(texto, largura=1200, altura=250):
    img = Image.new("RGB", (largura, altura), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    try:
        fonte = ImageFont.truetype("arial.ttf", 28)
    except Exception:
        fonte = ImageFont.load_default()

    cor_texto = (200, 0, 0)
    linhas = []
    linha_atual = ""

    for palavra in texto.split():
        teste = linha_atual + (" " if linha_atual else "") + palavra
        bbox = draw.textbbox((0, 0), teste, font=fonte)
        if bbox[2] - bbox[0] <= largura - 80:
            linha_atual = teste
        else:
            if linha_atual:
                linhas.append(linha_atual)
            linha_atual = palavra

    if linha_atual:
        linhas.append(linha_atual)

    alturas_linhas = []
    altura_total_texto = 0
    for linha in linhas:
        bbox = draw.textbbox((0, 0), linha, font=fonte)
        h = bbox[3] - bbox[1]
        alturas_linhas.append(h)
        altura_total_texto += h + 10

    y = (altura - altura_total_texto) // 2
    for i, linha in enumerate(linhas):
        bbox = draw.textbbox((0, 0), linha, font=fonte)
        x = (largura - (bbox[2] - bbox[0])) // 2
        draw.text((x, y), linha, fill=cor_texto, font=fonte)
        y += alturas_linhas[i] + 10

    return img


def processar_contas(base_dir: Path = Path(".")) -> dict:
    """
    Etapa 1: agrupa PDFs de contas com seus comprovantes e extrai dados de pagamento.

    Retorna um dicionário com:
        arquivos_gerados  — lista de Paths dos PDFs gerados em Agrupado/
        total_processadas — quantas contas foram processadas com sucesso
        total_erros       — quantas falharam
        arquivo_despesas  — Path do despesas_referencia.xlsx gerado
    """
    base_dir = Path(base_dir)

    pasta_contas = base_dir / "Contas"
    pasta_comprovantes = base_dir / "Comprovantes"
    pasta_saida = base_dir / "Agrupado"
    pasta_debug = base_dir / "debug_extracao"
    arquivo_despesas = base_dir / "Extratos" / "Entrada" / "despesas_referencia.xlsx"

    if not pasta_contas.exists():
        raise FileNotFoundError(f"Pasta não encontrada: {pasta_contas}")

    if not pasta_comprovantes.exists():
        raise FileNotFoundError(f"Pasta não encontrada: {pasta_comprovantes}")

    pasta_saida.mkdir(parents=True, exist_ok=True)

    arquivos_contas = sorted(f.name for f in pasta_contas.iterdir() if f.suffix.lower() == ".pdf")
    arquivos_comprovantes = sorted(f.name for f in pasta_comprovantes.iterdir() if f.suffix.lower() == ".pdf")

    print("=== CONTAS ENCONTRADAS ===")
    for f in arquivos_contas:
        print(f" - {f}")

    print("\n=== COMPROVANTES ENCONTRADOS ===")
    for f in arquivos_comprovantes:
        print(f" - {f}")

    comprovantes_dict = {
        obter_nome_base_comprovante(arq): arq
        for arq in arquivos_comprovantes
    }

    registros_despesas = []
    arquivos_gerados = []
    total_erros = 0

    print("\n=== PROCESSAMENTO ===")
    for arquivo_conta in arquivos_contas:
        nome_base = obter_nome_base_conta(arquivo_conta)
        sem_comprovante = nome_base not in comprovantes_dict

        caminho_conta = pasta_contas / arquivo_conta

        if sem_comprovante:
            print(f"Sem comprovante: {arquivo_conta}")
            arquivo_comprovante = None
            caminho_comprovante = None
        else:
            arquivo_comprovante = comprovantes_dict[nome_base]
            caminho_comprovante = pasta_comprovantes / arquivo_comprovante
            print(f"Processando: {arquivo_conta} + {arquivo_comprovante}")

        try:
            regra_conta = obter_regra_conta(nome_base)

            img_conta = pdf_para_imagem(
                caminho_conta, nome_base,
                pagina_idx=regra_conta["pagina"],
                y0=regra_conta["y0"],
                y1=regra_conta["y1"],
            )

            if regra_conta["redimensionar"] is not None:
                img_conta = redimensionar_imagem(img_conta, regra_conta["redimensionar"])

            if sem_comprovante:
                img_comprovante = criar_imagem_aviso(
                    "Pagamento por meio de débito automático. Validação por meio do extrato bancário",
                    largura=img_conta.width,
                    altura=180,
                )
            else:
                img_comprovante = pdf_para_imagem(
                    caminho_comprovante, nome_base,
                    pagina_idx=REGRA_COMPROVANTE["pagina"],
                    y0=REGRA_COMPROVANTE["y0"],
                    y1=REGRA_COMPROVANTE["y1"],
                )

            imagem_final = unir_imagens(img_conta, img_comprovante)

            caminho_saida = pasta_saida / f"{nome_base}.pdf"
            imagem_final.save(caminho_saida, "PDF")
            print(f"   Arquivo gerado: {caminho_saida}")
            arquivos_gerados.append(caminho_saida)

            dados_despesa = extrair_dados_despesa(
                caminho_conta=caminho_conta,
                caminho_comprovante=caminho_comprovante,
                nome_base=nome_base,
                pasta_debug=pasta_debug,
            )

            registros_despesas.append({
                "conta": nome_base.capitalize(),
                "arquivo_conta": arquivo_conta,
                "arquivo_comprovante": arquivo_comprovante or "",
                "arquivo_unificado": caminho_saida.name,
                "valor": dados_despesa["valor"],
                "data_pagamento": dados_despesa["data_pagamento"],
                "status_extracao": dados_despesa["status_extracao"],
            })

        except Exception as e:
            print(f"   Erro ao processar {arquivo_conta}: {e}")
            total_erros += 1
            continue

    if registros_despesas:
        arquivo_despesas.parent.mkdir(parents=True, exist_ok=True)
        pd.DataFrame(registros_despesas).to_excel(arquivo_despesas, index=False)
        print(f"\nPlanilha de despesas gerada em: {arquivo_despesas}")

    print(f"\nEtapa 1 finalizada! Processadas: {len(arquivos_gerados)} | Erros: {total_erros}")

    return {
        "arquivos_gerados": arquivos_gerados,
        "total_processadas": len(arquivos_gerados),
        "total_erros": total_erros,
        "arquivo_despesas": arquivo_despesas,
    }


# ==============================================================
# FUNÇÕES — ETAPA 2: CONFERÊNCIA DE EXTRATOS
# ==============================================================

def limpar_texto(texto: str) -> str:
    if texto is None:
        return ""
    return " ".join(texto.replace("\xa0", " ").split()).strip()


def extrair_texto_pagina(soup: BeautifulSoup) -> str:
    return limpar_texto(soup.get_text(" ", strip=True))


def localizar_metadado(texto_pagina: str, rotulo: str) -> str:
    padrao = rf"{re.escape(rotulo)}\s*:\s*(.+?)(?=\s+[A-ZÁÉÍÓÚÂÊÔÃÕÇ][^:]*:|$)"
    match = re.search(padrao, texto_pagina, flags=re.IGNORECASE)
    if match:
        return limpar_texto(match.group(1))
    return ""


def converter_valor_brasileiro(valor_str: str):
    texto = limpar_texto(valor_str)

    natureza = ""
    if texto.endswith(" D"):
        natureza = "D"
        texto = texto[:-2].strip()
    elif texto.endswith(" C"):
        natureza = "C"
        texto = texto[:-2].strip()

    texto = texto.replace(".", "").replace(",", ".")

    try:
        return float(texto), natureza
    except ValueError:
        return None, natureza


def encontrar_tabela_principal(soup: BeautifulSoup):
    for tabela in soup.find_all("table"):
        texto = limpar_texto(tabela.get_text(" ", strip=True)).lower()
        if all(k in texto for k in ("data mov", "nr. doc", "histórico", "valor", "saldo")):
            return tabela
    return None


def extrair_linhas_tabela(tabela):
    linhas = []
    for tr in tabela.find_all("tr"):
        colunas = tr.find_all(["td", "th"])
        valores = [limpar_texto(col.get_text(" ", strip=True)) for col in colunas]
        if any(valores):
            linhas.append(valores)
    return linhas


def padronizar_linhas_extrato(linhas_brutas, arquivo_origem, tipo_conta, metadados):
    registros = []
    ordem_linha = 0

    for linha in linhas_brutas:
        linha_lower = [item.lower() for item in linha]
        if "data mov." in linha_lower or "data mov" in linha_lower or \
           "nr. doc." in linha_lower or "nr. doc" in linha_lower:
            continue

        ordem_linha += 1

        data_mov = linha[0] if len(linha) > 0 else ""
        nr_doc = linha[1] if len(linha) > 1 else ""
        historico = linha[2] if len(linha) > 2 else ""
        valor_texto = linha[3] if len(linha) > 3 else ""
        saldo_texto = linha[4] if len(linha) > 4 else ""

        valor_num, natureza_valor = converter_valor_brasileiro(valor_texto)
        saldo_num, natureza_saldo = converter_valor_brasileiro(saldo_texto)

        historico_upper = historico.upper()
        if "SALDO ANTERIOR" in historico_upper:
            classe_linha = "saldo_anterior"
        elif "SALDO DIA" in historico_upper:
            classe_linha = "saldo_dia"
        else:
            classe_linha = "movimento"

        registros.append({
            "arquivo_origem": arquivo_origem,
            "tipo_conta": tipo_conta,
            "cliente": metadados.get("cliente", ""),
            "conta": metadados.get("conta", ""),
            "mes": metadados.get("mes", ""),
            "periodo": metadados.get("periodo", ""),
            "ordem_linha": ordem_linha,
            "classe_linha": classe_linha,
            "data_mov_texto": data_mov,
            "nr_doc": nr_doc,
            "historico": historico,
            "valor_texto": valor_texto,
            "valor_num": valor_num,
            "natureza_valor": natureza_valor,
            "saldo_texto": saldo_texto,
            "saldo_num": saldo_num,
            "natureza_saldo": natureza_saldo,
        })

    return registros


def ler_extrato_html(caminho_html: Path, tipo_conta: str) -> pd.DataFrame:
    if not caminho_html.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_html}")

    conteudo = None
    for cod in ["utf-8", "cp1252", "latin-1"]:
        try:
            with open(caminho_html, "r", encoding=cod) as f:
                conteudo = f.read()
            print(f"   Arquivo lido com codificação: {cod}")
            break
        except UnicodeDecodeError:
            continue

    if conteudo is None:
        raise RuntimeError(f"Não foi possível ler o arquivo {caminho_html}")

    soup = BeautifulSoup(conteudo, "lxml")
    texto_pagina = extrair_texto_pagina(soup)

    metadados = {
        "cliente": localizar_metadado(texto_pagina, "Cliente"),
        "conta": localizar_metadado(texto_pagina, "Conta"),
        "mes": localizar_metadado(texto_pagina, "Mês"),
        "periodo": localizar_metadado(texto_pagina, "Período"),
    }

    tabela_principal = encontrar_tabela_principal(soup)
    if tabela_principal is None:
        raise RuntimeError(f"Tabela principal não localizada em: {caminho_html}")

    registros = padronizar_linhas_extrato(
        linhas_brutas=extrair_linhas_tabela(tabela_principal),
        arquivo_origem=caminho_html.name,
        tipo_conta=tipo_conta,
        metadados=metadados,
    )

    df = pd.DataFrame(registros)
    df["data_mov"] = pd.to_datetime(df["data_mov_texto"], format="%d/%m/%Y", errors="coerce")
    return df


def ler_df_despesas_referencia(arquivo_despesas: Path) -> pd.DataFrame:
    if not arquivo_despesas.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {arquivo_despesas}")

    df = pd.read_excel(arquivo_despesas)

    faltando = {"conta", "valor", "data_pagamento"} - set(df.columns)
    if faltando:
        raise RuntimeError(f"Planilha de despesas sem colunas obrigatórias. Faltando: {sorted(faltando)}")

    df["conta"] = df["conta"].astype(str).str.strip()
    df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
    df["data_pagamento"] = pd.to_datetime(df["data_pagamento"], dayfirst=True, errors="coerce")

    df_validas = df[df["conta"].ne("") & df["valor"].notna() & df["data_pagamento"].notna()].copy()

    total = len(df)
    validas = len(df_validas)
    print(f"   Total de despesas: {total} | Válidas: {validas} | Pendentes: {total - validas}")

    if validas == 0:
        raise RuntimeError("Nenhuma despesa válida foi encontrada na planilha de referência.")

    return df_validas


def encontrar_matches(df_extratos: pd.DataFrame, df_despesas: pd.DataFrame) -> pd.DataFrame:
    registros_conferencia = []

    df_movimentos = df_extratos[
        (df_extratos["classe_linha"] == "movimento") &
        (df_extratos["natureza_valor"] == "D")
    ].copy()

    for _, despesa in df_despesas.iterrows():
        conta_nome = despesa["conta"]
        valor_esperado = despesa["valor"]
        data_esperada = despesa["data_pagamento"]

        candidatos = df_movimentos[
            (df_movimentos["valor_num"] == valor_esperado)
            & (df_movimentos["data_mov"] >= data_esperada - pd.Timedelta(days=3))
            & (df_movimentos["data_mov"] <= data_esperada + pd.Timedelta(days=3))
        ].copy()

        base = {
            "conta_mockada": conta_nome,
            "valor_esperado": valor_esperado,
            "data_esperada": data_esperada,
            "criterio_match": "valor_exato + data_proxima_3_dias",
        }

        if len(candidatos) == 0:
            registros_conferencia.append({
                **base,
                "status": "sem_match",
                "qtd_matches": 0,
                "arquivo_origem": "",
                "tipo_conta": "",
                "data_encontrada": pd.NaT,
                "historico_encontrado": "",
                "valor_encontrado": None,
                "ordem_linha": None,
                "observacao": "Nenhuma linha encontrada",
            })
            continue

        status = "match_forte" if len(candidatos) == 1 else "match_ambiguo"
        observacao = "Match único encontrado" if len(candidatos) == 1 else "Mais de uma linha encontrada para a mesma despesa"

        for _, candidato in candidatos.iterrows():
            registros_conferencia.append({
                **base,
                "status": status,
                "qtd_matches": len(candidatos),
                "arquivo_origem": candidato["arquivo_origem"],
                "tipo_conta": candidato["tipo_conta"],
                "data_encontrada": candidato["data_mov"],
                "historico_encontrado": candidato["historico"],
                "valor_encontrado": candidato["valor_num"],
                "ordem_linha": candidato["ordem_linha"],
                "observacao": observacao,
            })

    return pd.DataFrame(registros_conferencia)


def aplicar_marcacoes_no_extrato(df_extratos: pd.DataFrame, df_conferencia: pd.DataFrame) -> pd.DataFrame:
    df_saida = df_extratos.copy()
    df_saida["tem_match"] = ""
    df_saida["conta_associada"] = ""
    df_saida["status_match"] = ""
    df_saida["criterio_match"] = ""
    df_saida["observacao_match"] = ""

    for _, conf in df_conferencia.iterrows():
        arquivo_origem = conf["arquivo_origem"]
        ordem_linha = conf["ordem_linha"]

        if pd.isna(ordem_linha) or not arquivo_origem:
            continue

        filtro = (
            (df_saida["arquivo_origem"] == arquivo_origem) &
            (df_saida["ordem_linha"] == ordem_linha)
        )

        df_saida.loc[filtro, "tem_match"] = "sim"
        df_saida.loc[filtro, "conta_associada"] = conf["conta_mockada"]
        df_saida.loc[filtro, "status_match"] = conf["status"]
        df_saida.loc[filtro, "criterio_match"] = conf["criterio_match"]
        df_saida.loc[filtro, "observacao_match"] = conf["observacao"]

    return df_saida


def aplicar_formatacao_visual_excel(caminho_excel: Path):
    wb = load_workbook(caminho_excel)

    fill_verde = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
    fill_amarelo = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
    fill_vermelho = PatternFill(fill_type="solid", start_color="F4CCCC", end_color="F4CCCC")
    fill_azul_claro = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
    fonte_negrito = Font(bold=True)

    for nome_aba in ["extrato_corrente", "extrato_poupanca"]:
        if nome_aba not in wb.sheetnames:
            continue

        ws = wb[nome_aba]
        cabecalhos = {}
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=1, column=col).value
            if valor:
                cabecalhos[str(valor)] = col
                ws.cell(row=1, column=col).font = fonte_negrito

        col_tem_match = cabecalhos.get("tem_match")
        col_status_match = cabecalhos.get("status_match")

        if not col_tem_match:
            continue

        for row in range(2, ws.max_row + 1):
            tem_match = ws.cell(row=row, column=col_tem_match).value
            status_match = ws.cell(row=row, column=col_status_match).value if col_status_match else ""

            if tem_match == "sim":
                if status_match == "match_forte":
                    fill = fill_verde
                elif status_match == "match_ambiguo":
                    fill = fill_amarelo
                else:
                    fill = fill_azul_claro
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill

        ws.freeze_panes = "A2"

    if "conferencia" in wb.sheetnames:
        ws = wb["conferencia"]
        cabecalhos = {}
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=1, column=col).value
            if valor:
                cabecalhos[str(valor)] = col
                ws.cell(row=1, column=col).font = fonte_negrito

        col_status = cabecalhos.get("status")
        if col_status:
            mapa_fill = {
                "match_forte": fill_verde,
                "match_ambiguo": fill_amarelo,
                "sem_match": fill_vermelho,
            }
            for row in range(2, ws.max_row + 1):
                status = ws.cell(row=row, column=col_status).value
                if status in mapa_fill:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = mapa_fill[status]

        ws.freeze_panes = "A2"

    wb.save(caminho_excel)


def obter_estilo_html_por_status(status_match: str) -> dict:
    estilos = {
        "match_forte":  {"cor_borda": "#2E7D32", "cor_fundo": "#F3FBF4", "cor_texto": "#1B5E20"},
        "match_ambiguo": {"cor_borda": "#B26A00", "cor_fundo": "#FFF9ED", "cor_texto": "#8A5300"},
        "sem_match":    {"cor_borda": "#C62828", "cor_fundo": "#FFF4F4", "cor_texto": "#8E0000"},
    }
    return estilos.get(status_match, {"cor_borda": "#1565C0", "cor_fundo": "#F4F8FF", "cor_texto": "#0D47A1"})


def obter_estilo_html_saldo() -> dict:
    return {"cor_borda": "#1565C0", "cor_fundo": "#F4F8FF", "cor_texto": "#0D47A1"}


def extrair_mes_ano_do_texto_mes(texto_mes: str):
    mapa_meses = {
        "janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3,
        "abril": 4, "maio": 5, "junho": 6, "julho": 7,
        "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
    }

    if not texto_mes or "/" not in texto_mes:
        return None, None

    nome_mes, ano = texto_mes.strip().lower().split("/", 1)
    mes_num = mapa_meses.get(nome_mes.strip())
    if mes_num is None:
        return None, None

    try:
        return mes_num, int(ano.strip())
    except ValueError:
        return None, None


def aplicar_contorno_linha_html(tr, estilo_dict):
    cor_borda = estilo_dict["cor_borda"]
    cor_fundo = estilo_dict["cor_fundo"]

    tds = tr.find_all("td")
    if not tds:
        return

    for i, td in enumerate(tds):
        estilos = [
            f"background-color: {cor_fundo}",
            f"border-top: 2px solid {cor_borda}",
            f"border-bottom: 2px solid {cor_borda}",
        ]
        if i == 0:
            estilos.append(f"border-left: 2px solid {cor_borda}")
        if i == len(tds) - 1:
            estilos.append(f"border-right: 2px solid {cor_borda}")

        existente = td.get("style", "")
        novo = "; ".join(estilos)
        td["style"] = (existente.rstrip(";") + "; " + novo) if existente else novo


def gerar_html_destacado(
    caminho_html_entrada: Path,
    caminho_html_saida: Path,
    df_extrato_marcado: pd.DataFrame,
):
    conteudo = None
    for cod in ["utf-8", "cp1252", "latin-1"]:
        try:
            with open(caminho_html_entrada, "r", encoding=cod) as f:
                conteudo = f.read()
            print(f"   HTML base lido com codificação: {cod}")
            break
        except UnicodeDecodeError:
            continue

    if conteudo is None:
        raise RuntimeError(f"Não foi possível ler o HTML {caminho_html_entrada}")

    soup = BeautifulSoup(conteudo, "lxml")

    tabela_principal = encontrar_tabela_principal(soup)
    if tabela_principal is None:
        raise RuntimeError(f"Tabela principal não localizada em: {caminho_html_entrada}")

    df_matches = df_extrato_marcado[df_extrato_marcado["tem_match"] == "sim"].copy()
    mapa_matches = {
        row["ordem_linha"]: {
            "status_match": row.get("status_match", ""),
            "conta_associada": row.get("conta_associada", ""),
            "observacao_match": row.get("observacao_match", ""),
        }
        for _, row in df_matches.iterrows()
    }

    ordens_saldo_anterior = set()
    ordem_ultimo_saldo_dia_mes = None

    if not df_extrato_marcado.empty:
        df_sa = df_extrato_marcado[df_extrato_marcado["classe_linha"] == "saldo_anterior"]
        ordens_saldo_anterior = set(df_sa["ordem_linha"].dropna().astype(int).tolist())

        texto_mes = ""
        if "mes" in df_extrato_marcado.columns and not df_extrato_marcado["mes"].dropna().empty:
            texto_mes = str(df_extrato_marcado["mes"].dropna().iloc[0])

        mes_ref, ano_ref = extrair_mes_ano_do_texto_mes(texto_mes)
        df_sd = df_extrato_marcado[df_extrato_marcado["classe_linha"] == "saldo_dia"].copy()

        if mes_ref and ano_ref and "data_mov" in df_sd.columns:
            df_sd_mes = df_sd[
                (df_sd["data_mov"].dt.month == mes_ref) &
                (df_sd["data_mov"].dt.year == ano_ref)
            ]
            if not df_sd_mes.empty:
                ordem_ultimo_saldo_dia_mes = int(df_sd_mes["ordem_linha"].max())

    ordem_linha = 0
    for tr in tabela_principal.find_all("tr"):
        colunas = tr.find_all(["td", "th"])
        valores = [limpar_texto(col.get_text(" ", strip=True)) for col in colunas]

        if not any(valores):
            continue

        linha_lower = [v.lower() for v in valores]
        if "data mov." in linha_lower or "data mov" in linha_lower or \
           "nr. doc." in linha_lower or "nr. doc" in linha_lower:
            continue

        ordem_linha += 1

        if ordem_linha in ordens_saldo_anterior:
            aplicar_contorno_linha_html(tr, obter_estilo_html_saldo())
            tr["title"] = "Saldo anterior"

        if ordem_ultimo_saldo_dia_mes is not None and ordem_linha == ordem_ultimo_saldo_dia_mes:
            aplicar_contorno_linha_html(tr, obter_estilo_html_saldo())
            tr["title"] = "Último saldo dia do mês do extrato"

        if ordem_linha in mapa_matches:
            info = mapa_matches[ordem_linha]
            aplicar_contorno_linha_html(tr, obter_estilo_html_por_status(info["status_match"]))

            tooltip = f"{info['conta_associada']} | {info['status_match']}"
            if info["observacao_match"]:
                tooltip += f" | {info['observacao_match']}"
            tr["title"] = tooltip

            tds = tr.find_all("td")
            if len(tds) >= 3:
                td_historico = tds[2]

                wrapper = soup.new_tag("div")
                wrapper["style"] = (
                    "display: flex; justify-content: space-between; "
                    "align-items: center; gap: 12px; width: 100%;"
                )

                span_texto = soup.new_tag("span")
                span_texto.string = limpar_texto(td_historico.get_text(" ", strip=True))

                span_badge = soup.new_tag("span")
                span_badge["style"] = (
                    "display: inline-block; padding: 3px 8px; border-radius: 10px; "
                    "border: 1px solid #1F4E78; background-color: #FFFFFF; color: #1F4E78; "
                    "font-size: 11px; font-weight: bold; white-space: nowrap;"
                )
                span_badge.string = str(info["conta_associada"]).upper()

                td_historico.clear()
                wrapper.append(span_texto)
                wrapper.append(span_badge)
                td_historico.append(wrapper)

    if soup.head is None:
        head_tag = soup.new_tag("head")
        soup.html.insert(0, head_tag) if soup.html else soup.insert(0, head_tag)

    style_tag = soup.new_tag("style")
    style_tag.string = """
    @media print {
        * { -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }
        table { border-collapse: collapse !important; }
        tr, td { page-break-inside: avoid !important; }
        td { box-sizing: border-box !important; }
    }
    """
    soup.head.append(style_tag)

    with open(caminho_html_saida, "w", encoding="utf-8") as f:
        f.write(str(soup))

    print(f"   HTML destacado gerado: {caminho_html_saida}")


def exportar_htmls_destacados(
    df_corrente_marcado: pd.DataFrame,
    df_poupanca_marcado: pd.DataFrame,
    arquivo_extrato_corrente: Path,
    arquivo_extrato_poupanca: Path,
    arquivo_html_corrente: Path,
    arquivo_html_poupanca: Path,
):
    print("\nGerando HTML destacado do extrato corrente...")
    gerar_html_destacado(arquivo_extrato_corrente, arquivo_html_corrente, df_corrente_marcado)

    print("\nGerando HTML destacado do extrato poupança...")
    gerar_html_destacado(arquivo_extrato_poupanca, arquivo_html_poupanca, df_poupanca_marcado)


def exportar_excel(
    df_corrente_marcado: pd.DataFrame,
    df_poupanca_marcado: pd.DataFrame,
    df_conferencia: pd.DataFrame,
    arquivo_saida_excel: Path,
):
    arquivo_saida_excel.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(arquivo_saida_excel, engine="openpyxl") as writer:
        df_corrente_marcado.to_excel(writer, sheet_name="extrato_corrente", index=False)
        df_poupanca_marcado.to_excel(writer, sheet_name="extrato_poupanca", index=False)
        df_conferencia.to_excel(writer, sheet_name="conferencia", index=False)

    aplicar_formatacao_visual_excel(arquivo_saida_excel)
    print(f"\nArquivo Excel gerado: {arquivo_saida_excel}")


def processar_extratos(base_dir: Path = Path(".")) -> dict:
    """
    Etapa 2: lê extratos bancários HTML e confere cada despesa da planilha de referência.

    Retorna um dicionário com:
        total_despesas       — total de despesas carregadas
        match_forte          — despesas com match único no extrato
        match_ambiguo        — despesas com mais de um candidato
        sem_match            — despesas sem nenhum match
        arquivo_excel        — Path do conferencia_extratos.xlsx
        arquivo_html_corrente — Path do extrato corrente destacado
        arquivo_html_poupanca — Path do extrato poupança destacado
    """
    base_dir = Path(base_dir)

    pasta_entrada = base_dir / "Extratos" / "Entrada"
    pasta_saida = base_dir / "Extratos" / "Saída"

    arquivo_extrato_corrente = pasta_entrada / "Extrato Corrente.html"
    arquivo_extrato_poupanca = pasta_entrada / "Extrato Poupança.html"
    arquivo_despesas = pasta_entrada / "despesas_referencia.xlsx"
    arquivo_saida_excel = pasta_saida / "conferencia_extratos.xlsx"
    arquivo_html_corrente = pasta_saida / "extrato_corrente_destacado.html"
    arquivo_html_poupanca = pasta_saida / "extrato_poupanca_destacado.html"

    pasta_saida.mkdir(parents=True, exist_ok=True)

    print("=== INÍCIO DO PROCESSAMENTO DOS EXTRATOS ===")

    print(f"\nLendo extrato corrente: {arquivo_extrato_corrente.name}")
    df_corrente = ler_extrato_html(arquivo_extrato_corrente, "corrente")
    print(f"   OK - {len(df_corrente)} linhas extraídas")

    print(f"\nLendo extrato poupança: {arquivo_extrato_poupanca.name}")
    df_poupanca = ler_extrato_html(arquivo_extrato_poupanca, "poupanca")
    print(f"   OK - {len(df_poupanca)} linhas extraídas")

    df_extratos = pd.concat([df_corrente, df_poupanca], ignore_index=True)
    print(f"\nTotal consolidado: {len(df_extratos)} linhas")

    df_despesas = ler_df_despesas_referencia(arquivo_despesas)
    print(f"Despesas carregadas para conferência: {len(df_despesas)}")

    df_conferencia = encontrar_matches(df_extratos, df_despesas)
    print(f"Registros de conferência gerados: {len(df_conferencia)}")

    df_corrente_marcado = aplicar_marcacoes_no_extrato(df_corrente, df_conferencia)
    df_poupanca_marcado = aplicar_marcacoes_no_extrato(df_poupanca, df_conferencia)

    exportar_excel(df_corrente_marcado, df_poupanca_marcado, df_conferencia, arquivo_saida_excel)

    exportar_htmls_destacados(
        df_corrente_marcado, df_poupanca_marcado,
        arquivo_extrato_corrente, arquivo_extrato_poupanca,
        arquivo_html_corrente, arquivo_html_poupanca,
    )

    contagem = df_conferencia["status"].value_counts().to_dict() if not df_conferencia.empty else {}

    print(
        f"\nEtapa 2 finalizada! "
        f"match_forte={contagem.get('match_forte', 0)} | "
        f"match_ambiguo={contagem.get('match_ambiguo', 0)} | "
        f"sem_match={contagem.get('sem_match', 0)}"
    )

    return {
        "total_despesas": len(df_despesas),
        "match_forte": contagem.get("match_forte", 0),
        "match_ambiguo": contagem.get("match_ambiguo", 0),
        "sem_match": contagem.get("sem_match", 0),
        "arquivo_excel": arquivo_saida_excel,
        "arquivo_html_corrente": arquivo_html_corrente,
        "arquivo_html_poupanca": arquivo_html_poupanca,
    }


# ==============================================================
# EXECUÇÃO PRINCIPAL
# ==============================================================
if __name__ == "__main__":
    base = Path(".")

    print("=" * 60)
    print("ETAPA 1 — AGRUPAMENTO DE CONTAS E COMPROVANTES")
    print("=" * 60)
    processar_contas(base)

    print("\n" + "=" * 60)
    print("ETAPA 2 — CONFERÊNCIA DOS EXTRATOS BANCÁRIOS")
    print("=" * 60)
    processar_extratos(base)