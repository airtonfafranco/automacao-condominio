#!/usr/bin/env python3
"""
gerar_balancete.py — Etapa 3 da automação do Condomínio Barão de Mesquita.

Consome:
  - config_balancete.json              (mesmo diretório deste script)
  - Extratos/Entrada/Extrato Corrente.html
  - Extratos/Entrada/Extrato Poupança.html
  - Extratos/Entrada/despesas_referencia.xlsx  (saída Etapa 1)
  - Planilha TACON_TUV.xlsx            (passada como parâmetro)
  - inadimplentes JSON do mês anterior (passado como parâmetro ou buscado em base_dir)

Gera:
  - balancete_{mmmANO}.xlsx
  - inadimplentes_{mmmANO}.json
"""

import json
import calendar
from datetime import date
from pathlib import Path
from typing import Optional

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.cell import range_boundaries


# ═══════════════════════════════════════════════════════════════════
# CONSTANTES
# ═══════════════════════════════════════════════════════════════════

_DIR = Path(__file__).parent
CONFIG_PATH = _DIR / "config_balancete.json"

MESES_PT = {
    1: "janeiro",  2: "fevereiro", 3: "março",    4: "abril",
    5: "maio",     6: "junho",     7: "julho",     8: "agosto",
    9: "setembro", 10: "outubro",  11: "novembro", 12: "dezembro",
}

# (mes_anterior, delta_ano) para cada mês
MES_ANT = {
    1: (12, -1), 2: (1, 0),  3: (2, 0),  4: (3, 0),
    5: (4, 0),   6: (5, 0),  7: (6, 0),  8: (7, 0),
    9: (8, 0),   10: (9, 0), 11: (10, 0),12: (11, 0),
}


# ═══════════════════════════════════════════════════════════════════
# PARSING DE EXTRATOS HTML
# ═══════════════════════════════════════════════════════════════════

def _limpar(t: str) -> str:
    if t is None:
        return ""
    return " ".join(t.replace("\xa0", " ").split()).strip()


def _conv_valor(texto: str):
    """Converte '1.234,56 D' → (1234.56, 'D')"""
    t = _limpar(texto)
    nat = ""
    if t.endswith(" D"):
        nat, t = "D", t[:-2].strip()
    elif t.endswith(" C"):
        nat, t = "C", t[:-2].strip()
    t = t.replace(".", "").replace(",", ".")
    try:
        return float(t), nat
    except ValueError:
        return None, nat


def _achar_tabela(soup: BeautifulSoup):
    for tab in soup.find_all("table"):
        txt = _limpar(tab.get_text(" ", strip=True)).lower()
        if all(k in txt for k in ("data mov", "histórico", "valor", "saldo")):
            return tab
    return None


def ler_extrato(caminho: Path, tipo: str) -> list:
    """Lê extrato HTML e retorna lista de dicts com os lançamentos."""
    conteudo = None
    for cod in ["utf-8", "cp1252", "latin-1"]:
        try:
            conteudo = caminho.read_text(encoding=cod)
            break
        except UnicodeDecodeError:
            continue
    if conteudo is None:
        raise RuntimeError(f"Não foi possível ler: {caminho}")

    soup = BeautifulSoup(conteudo, "lxml")
    tab = _achar_tabela(soup)
    if tab is None:
        raise RuntimeError(f"Tabela principal não encontrada em: {caminho}")

    rows = []
    for tr in tab.find_all("tr"):
        cols = [_limpar(td.get_text(" ", strip=True)) for td in tr.find_all(["td", "th"])]
        if not any(cols):
            continue
        lower_join = " ".join(c.lower() for c in cols)
        if any(k in lower_join for k in ["data mov.", "nr. doc", "histórico"]):
            continue
        if len(cols) < 4:
            continue

        hist = cols[2] if len(cols) > 2 else ""
        val, nat = _conv_valor(cols[3] if len(cols) > 3 else "")
        saldo, _ = _conv_valor(cols[4] if len(cols) > 4 else "")
        hu = hist.upper()

        if "SALDO ANTERIOR" in hu:
            cls = "saldo_anterior"
        elif "SALDO DIA" in hu:
            cls = "saldo_dia"
        else:
            cls = "mov"

        rows.append({
            "tipo":  tipo,
            "data":  cols[0],
            "hist":  hist,
            "val":   val,
            "nat":   nat,
            "saldo": saldo,
            "cls":   cls,
        })
    return rows


def _parse_data(texto: str, ano: int) -> Optional[date]:
    p = texto.strip().split("/")
    try:
        if len(p) == 2:
            return date(ano, int(p[1]), int(p[0]))
        if len(p) == 3:
            return date(int(p[2]), int(p[1]), int(p[0]))
    except (ValueError, IndexError):
        pass
    return None


def _movs_do_mes(rows: list, mes: int, ano: int) -> list:
    """Filtra lançamentos do mês de referência."""
    return [
        r for r in rows
        if r["cls"] == "mov"
        and (d := _parse_data(r["data"], ano)) is not None
        and d.month == mes and d.year == ano
    ]


# ═══════════════════════════════════════════════════════════════════
# CÁLCULO DE SALDOS
# ═══════════════════════════════════════════════════════════════════

def calcular_saldos_cc(rows: list, mes: int, ano: int) -> tuple:
    """Saldo anterior CC e saldo final CC (último SALDO DIA do mês)."""
    movs = [r for r in rows if r["cls"] == "mov"]
    dias = [r for r in rows if r["cls"] == "saldo_dia"]

    anterior = None
    if movs and movs[0]["val"] is not None and movs[0]["saldo"] is not None:
        m = movs[0]
        anterior = round(m["saldo"] + m["val"], 2) if m["nat"] == "D" \
               else round(m["saldo"] - m["val"], 2)

    final = None
    for r in reversed(dias):
        d = _parse_data(r["data"], ano)
        if d and d.month == mes and d.year == ano:
            final = r["saldo"]
            break

    return anterior, final


def calcular_saldos_poupanca(rows: list, mes: int, ano: int) -> tuple:
    """Saldo anterior Poupança e saldo final Poupança."""
    movs = [r for r in rows if r["cls"] == "mov"]
    dias = [r for r in rows if r["cls"] == "saldo_dia"]

    anterior = None
    if movs and movs[0]["val"] is not None and movs[0]["saldo"] is not None:
        m = movs[0]
        anterior = round(m["saldo"] - m["val"], 2) if m["nat"] == "C" \
               else round(m["saldo"] + m["val"], 2)

    final = None
    for r in reversed(dias):
        d = _parse_data(r["data"], ano)
        if d and d.month == mes and d.year == ano:
            final = r["saldo"]
            break

    return anterior, final


# ═══════════════════════════════════════════════════════════════════
# RECEITAS E DESPESAS DO EXTRATO
# ═══════════════════════════════════════════════════════════════════

def calcular_poupanca(rows: list, config: dict, mes: int, ano: int) -> dict:
    """Juros recebidos (crédito) e IRRF pago (débito) na poupança."""
    cred_hist = [h.upper() for h in config["poupanca"]["credito_historicos"]]
    deb_hist  = [h.upper() for h in config["poupanca"]["debito_historicos"]]
    movs = _movs_do_mes(rows, mes, ano)

    juros = 0.0
    irrf  = 0.0
    for r in movs:
        if not r["val"]:
            continue
        hu = r["hist"].upper()
        if r["nat"] == "C" and any(h in hu for h in cred_hist):
            juros += r["val"]
        elif r["nat"] == "D" and any(h in hu for h in deb_hist):
            irrf  += r["val"]

    return {"juros": round(juros, 2), "irrf": round(irrf, 2)}


def calcular_despesas_bancarias_cc(rows: list, config: dict, mes: int, ano: int) -> float:
    """Soma das tarifas bancárias da conta corrente."""
    alvo = [h.upper() for h in config["despesas_bancarias"]["cc_historicos"]]
    movs = _movs_do_mes(rows, mes, ano)

    total = 0.0
    for r in movs:
        if r["nat"] != "D" or not r["val"]:
            continue
        hu = r["hist"].upper()
        if any(hu.startswith(h) or h in hu for h in alvo):
            total += r["val"]
    return round(total, 2)


def calcular_cobs_e_extras(rows: list, config: dict, mes: int, ano: int) -> dict:
    """Extrai COBs (pagamentos dos condôminos) e receitas extras (CRED TEV)."""
    extras_hist = [h.upper() for h in config["receitas_extras"]["historicos"]]
    movs = _movs_do_mes(rows, mes, ano)

    cob_values      = []
    receitas_extras = []
    for r in movs:
        if r["nat"] != "C" or not r["val"]:
            continue
        hu = r["hist"].upper()
        if hu.startswith("COB"):
            cob_values.append(r["val"])
        elif any(h in hu for h in extras_hist):
            d = _parse_data(r["data"], ano)
            receitas_extras.append({"data": d, "hist": r["hist"], "val": r["val"]})

    return {"cob_values": cob_values, "receitas_extras": receitas_extras}


# ═══════════════════════════════════════════════════════════════════
# ALGORITMO TACON / TUV / MULTA / REF ANTERIOR
# ═══════════════════════════════════════════════════════════════════

def calcular_tacon_tuv(
    cob_values:             list,
    ref_anterior_principal: float,
    tacon_esperado:         float,
    tuv_esperado:           float,
    df_tacon_tuv:           pd.DataFrame,
    inadimplente_forcado:   Optional[dict] = None,
) -> dict:
    """
    Calcula TACON, TUV, Multa e Referência mês anterior a partir dos COBs.

    Lógica:
      1. total_cob = soma de todos os COBs
      2. Juros = parte não-múltipla de R$100 de cada COB
      3. tacon_tuv_mes = total_cob - total_juros - ref_anterior_principal
      4. COBs com parte inteira > 300 carregam juro do mês anterior
      5. multa = total_juros - juro_ref_anterior
      6. shortfall = esperado - recebido → identifica inadimplente

    inadimplente_forcado: linha do df_tacon_tuv já resolvida manualmente (resolve ambiguidade).
    """
    total_cob = round(sum(cob_values), 2)

    # Juros = resto da divisão por 100 de cada COB
    juros_por_cob: dict[float, float] = {}
    for v in cob_values:
        rem = round(v % 100, 2)
        if rem > 0.01:
            juros_por_cob[v] = rem

    total_juros_all = round(sum(juros_por_cob.values()), 2)

    # TACON+TUV efetivamente arrecadados no mês
    tacon_tuv_mes = round(total_cob - total_juros_all - ref_anterior_principal, 2)

    # Separar juro do mês anterior do juro de mora do mês corrente
    ref_anterior_juros = 0.0
    if ref_anterior_principal > 0:
        for v, rem in juros_por_cob.items():
            parte_inteira = round(v - rem, 2)
            if parte_inteira > 300:
                ref_anterior_juros += rem
        ref_anterior_juros = round(ref_anterior_juros, 2)

    multa         = round(total_juros_all - ref_anterior_juros, 2)
    ref_ant_total = round(ref_anterior_principal + ref_anterior_juros, 2)

    # ── Identificar inadimplente ──
    total_esperado   = round(tacon_esperado + tuv_esperado, 2)
    shortfall        = round(total_esperado - tacon_tuv_mes, 2)
    inadimplente_mes = None
    ambiguidade      = False
    candidatos       = []   # lista de dicts para exibir na UI

    if shortfall > 0.01:
        # Resolve direto se vier forçado (UI confirmou manualmente)
        if inadimplente_forcado is not None:
            inadimplente_mes = inadimplente_forcado
            ambiguidade      = False
            print(f"  ✅ Inadimplente confirmado manualmente: "
                  f"{inadimplente_forcado.get('unidade', '?')} "
                  f"(shortfall R$ {shortfall:.2f})")
        else:
            # Busca automática na tabela
            col_unit  = None
            col_total = None
            col_tacon = None
            col_tuv   = None

            for c in df_tacon_tuv.columns:
                cu = c.strip().upper()
                if cu == "UNIDADE":
                    col_unit = c
                elif cu in ("TACON+TUV", "TOTAL"):
                    col_total = c
                elif cu == "TACON":
                    col_tacon = c
                elif cu == "TUV":
                    col_tuv = c

            if col_total is None and col_tacon and col_tuv:
                df_tacon_tuv = df_tacon_tuv.copy()
                df_tacon_tuv["_TOTAL_"] = (
                    pd.to_numeric(df_tacon_tuv[col_tacon], errors="coerce").fillna(0) +
                    pd.to_numeric(df_tacon_tuv[col_tuv],   errors="coerce").fillna(0)
                )
                col_total = "_TOTAL_"

            if col_total:
                col_num = pd.to_numeric(df_tacon_tuv[col_total], errors="coerce")
                matches = df_tacon_tuv[col_num.apply(lambda x: abs(x - shortfall) < 0.01)]

                if len(matches) == 1:
                    inadimplente_mes = matches.iloc[0].to_dict()
                    ambiguidade      = False
                    print(f"  ✅ Inadimplente identificado automaticamente: "
                          f"{inadimplente_mes.get(col_unit, '?')} "
                          f"(shortfall R$ {shortfall:.2f})")
                elif len(matches) > 1:
                    candidatos       = matches.to_dict(orient="records")
                    inadimplente_mes = candidatos
                    ambiguidade      = True
                    print(f"  ⚠️  Ambiguidade: shortfall R$ {shortfall:.2f} → "
                          f"{len(matches)} unidades possíveis: "
                          f"{[m.get(col_unit, '?') for m in candidatos]}")
                    print("     Confirme manualmente qual unidade não pagou.")
                else:
                    ambiguidade = True
                    print(f"  ⚠️  Nenhuma unidade com total = R$ {shortfall:.2f}. "
                          f"Verifique manualmente.")

    # ── Calcular TACON e TUV finais (descontando inadimplente não ambíguo) ──
    col_tacon_val = 0.0
    col_tuv_val   = 0.0
    if shortfall > 0.01 and not ambiguidade and isinstance(inadimplente_mes, dict):
        for c in df_tacon_tuv.columns:
            cu = c.strip().upper()
            if cu == "TACON":
                col_tacon_val = float(inadimplente_mes.get(c, 0) or 0)
            if cu == "TUV":
                col_tuv_val   = float(inadimplente_mes.get(c, 0) or 0)

    tacon_final = round(tacon_esperado - col_tacon_val, 2)
    tuv_final   = round(tuv_esperado   - col_tuv_val,   2)

    return {
        "total_cob":          total_cob,
        "tacon_tuv_mes":      tacon_tuv_mes,
        "tacon":              tacon_final,
        "tuv":                tuv_final,
        "multa":              multa,
        "ref_anterior_total": ref_ant_total,
        "shortfall":          shortfall,
        "inadimplente_mes":   inadimplente_mes,
        "ambiguidade":        ambiguidade,
        "candidatos":         candidatos,
    }


# ═══════════════════════════════════════════════════════════════════
# LEITURA DE ARQUIVOS DE ENTRADA
# ═══════════════════════════════════════════════════════════════════

def ler_config() -> dict:
    """Carrega config_balancete.json do mesmo diretório do script."""
    if not CONFIG_PATH.exists():
        raise FileNotFoundError(f"Arquivo de configuração não encontrado: {CONFIG_PATH}")
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


def ler_inadimplentes_anteriores(
    base_dir: Path,
    mes_ant: int,
    ano_ant: int,
    dados_externos: Optional[dict] = None,
) -> dict:
    """
    Retorna inadimplentes do mês anterior.
    Prioridade: dados_externos (upload) → arquivo JSON em base_dir.
    """
    if dados_externos is not None:
        print(f"  ✅ Inadimplentes anteriores carregados via upload.")
        return dados_externos

    nome_mes = MESES_PT[mes_ant]
    candidatos = [
        base_dir / f"inadimplentes_{nome_mes}{ano_ant}.json",
        base_dir / f"inadimplentes_{nome_mes}_{ano_ant}.json",
        base_dir / f"inadimplentes_{nome_mes[:3]}{ano_ant}.json",
    ]
    for c in candidatos:
        if c.exists():
            with open(c, encoding="utf-8") as f:
                return json.load(f)

    print(f"  ℹ️  Nenhum arquivo de inadimplentes encontrado para "
          f"{nome_mes}/{ano_ant}. Assumindo sem inadimplência anterior.")
    return {"competencia": f"{ano_ant}-{mes_ant:02d}", "inadimplentes": []}


def ler_tacon_tuv(caminho: Path) -> tuple:
    """
    Lê a planilha TACON_TUV e retorna (df, tacon_total, tuv_total).
    A primeira coluna é sempre renomeada para 'unidade'.
    """
    df = pd.read_excel(caminho)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.rename(columns={df.columns[0]: "unidade"})

    col_tacon = next((c for c in df.columns if c.upper() == "TACON"), None)
    col_tuv   = next((c for c in df.columns if c.upper() == "TUV"),   None)

    if col_tacon is None or col_tuv is None:
        raise ValueError(
            f"Colunas TACON/TUV não encontradas em {caminho}. "
            f"Colunas disponíveis: {list(df.columns)}"
        )

    tacon_total = float(pd.to_numeric(df[col_tacon], errors="coerce").fillna(0).sum())
    tuv_total   = float(pd.to_numeric(df[col_tuv],   errors="coerce").fillna(0).sum())

    return df, tacon_total, tuv_total


def ler_despesas_referencia(caminho: Path, config: dict) -> dict:
    """
    Lê despesas_referencia.xlsx e classifica cada conta:
      - "correntes": contas em CONTAS_CORRENTES
      - "servicos":  qualquer outra conta (inclusive não mapeadas)

    Retorna:
      { correntes: [...], servicos: [...], nao_mapeadas: [...] }
    """
    df = pd.read_excel(caminho)
    df.columns = [str(c).strip().lower() for c in df.columns]

    contas_config = {k.lower(): v for k, v in config["contas"].items()}

    correntes    = []
    servicos     = []
    nao_mapeadas = []

    for _, row in df.iterrows():
        nome_raw  = str(row.get("conta", "")).strip()
        nome_norm = nome_raw.lower()
        valor     = row.get("valor", None)
        data_txt  = str(row.get("data_pagamento", "")).strip()

        if valor is None or str(valor).strip() in ("", "nan"):
            continue
        try:
            valor = float(str(valor).replace(",", "."))
        except ValueError:
            continue

        # Parsear data
        from datetime import datetime as _dt
        data_obj = None
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                data_obj = _dt.strptime(data_txt, fmt).date()
                break
            except ValueError:
                continue

        # Busca exata, depois parcial
        cfg = contas_config.get(nome_norm)
        if cfg is None:
            for chave, vcfg in contas_config.items():
                if chave in nome_norm or nome_norm in chave:
                    cfg = vcfg
                    break

        data_fmt = data_obj.strftime("%d/%m/%Y") if data_obj else data_txt

        if cfg:
            entrada = {
                "descricao": f"{cfg['descricao']} ({data_fmt})",
                "valor":     valor,
                "data":      data_obj,
            }
            if cfg["categoria"] == "correntes":
                correntes.append(entrada)
            else:
                servicos.append(entrada)
        else:
            # Conta não mapeada → Serviço (regra V2)
            print(f"  ⚠️  Conta '{nome_raw}' não está no config. "
                  f"Classificada automaticamente como 'Serviços'.")
            nao_mapeadas.append(nome_norm)
            servicos.append({
                "descricao": f"{nome_raw.capitalize()} ({data_fmt})",
                "valor":     valor,
                "data":      data_obj,
            })

    def _sort_key(x):
        return x["data"] if x["data"] else date(9999, 12, 31)

    correntes.sort(key=_sort_key)
    servicos.sort(key=_sort_key)

    return {"correntes": correntes, "servicos": servicos, "nao_mapeadas": nao_mapeadas}


# ═══════════════════════════════════════════════════════════════════
# GERAÇÃO DO EXCEL — DINÂMICA (sem linhas fixas hardcoded)
# ═══════════════════════════════════════════════════════════════════

_THIN   = Side(style="thin")
_DOUBLE = Side(style="double")
_YELLOW = PatternFill("solid", fgColor="FFFF00")
_BOLD   = Font(bold=True, name="Calibri", size=10)
_NORM   = Font(name="Calibri", size=10)
_CRIGHT = Alignment(horizontal="right",  vertical="center")
_CCENT  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CLEFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_FMT_M  = "#,##0.00"

_B_THIN   = Border(left=_THIN,   right=_THIN,   top=_THIN,   bottom=_THIN)
_B_DOUBLE = Border(left=_DOUBLE, right=_DOUBLE, top=_DOUBLE, bottom=_DOUBLE)


def _mw(ws, rng: str, val, font=None, fill=None, border=None, align=None, num_fmt=None):
    """Mescla intervalo, escreve valor e aplica estilos."""
    ws.merge_cells(rng)
    min_col, min_row, max_col, max_row = range_boundaries(rng)
    cell = ws.cell(row=min_row, column=min_col)
    cell.value = val
    if font:    cell.font         = font
    if fill:    cell.fill         = fill
    if align:   cell.alignment    = align
    if num_fmt: cell.number_format = num_fmt
    if border:
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c).border = border
    return cell


def _w(ws, row, col, val, font=None, fill=None, border=None, align=None, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=val)
    if font:    cell.font          = font
    if fill:    cell.fill          = fill
    if align:   cell.alignment     = align
    if num_fmt: cell.number_format = num_fmt
    if border:  cell.border        = border
    return cell


def _linha_item(ws, row: int, num, descricao: str, valor: float):
    """Linha de item (despesa ou receita): nº | descrição | valor (C:E mesclado)."""
    _w(ws, row, 1, num,       font=_NORM, border=_B_THIN, align=_CCENT)
    _w(ws, row, 2, descricao, font=_NORM, border=_B_THIN, align=_CLEFT)
    ws.merge_cells(f"C{row}:E{row}")
    c = ws.cell(row=row, column=3, value=valor)
    c.font = _BOLD; c.number_format = _FMT_M
    c.alignment = _CRIGHT; c.border = _B_THIN
    for col in [4, 5]:
        ws.cell(row=row, column=col).border = _B_THIN


def _set_total_formula(ws, row_header: int, row_start: int, row_end: int, border=None):
    """Aplica fórmula de soma na célula C do cabeçalho da seção."""
    ws.merge_cells(f"C{row_header}:E{row_header}")
    c = ws.cell(row=row_header, column=3)
    c.value = f"=SUM(C{row_start}:C{row_end})"
    c.font = _BOLD
    c.number_format = _FMT_M
    c.alignment = _CRIGHT
    brd = border or _B_DOUBLE
    for col in [3, 4, 5]:
        ws.cell(row=row_header, column=col).border = brd


def gerar_excel_balancete(dados: dict, mes_nome: str, mes_num: int, ano: int,
                          output_path: Path):
    """
    Gera o XLSX do balancete com layout dinâmico.

    dados deve conter:
      correntes, servicos, receitas,
      saldo_ant_cc, saldo_ant_poup, caixa_ant,
      saldo_fin_cc, saldo_fin_poup, caixa_fin,
      data_ant_str, data_fin_str
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "RESUMO MES"

    for col, width in zip("ABCDE", [5.86, 52.71, 11.00, 10.43, 14.14]):
        ws.column_dimensions[col].width = width

    mes_upper = mes_nome.upper()
    r = 1

    # ── Cabeçalho ──
    _mw(ws, f"A{r}:E{r}",
        "CONDOMÍNIO BARÃO DE MESQUITA — CNPJ 00.978.554/0001-68",
        font=Font(bold=True, name="Calibri", size=12), align=_CCENT)
    r += 1

    _mw(ws, f"A{r}:E{r}",
        "Relatório Financeiro e Administrativo",
        font=Font(name="Calibri", size=11), align=_CCENT)
    r += 2  # +1 spacer

    # ── Título do balancete ──
    _mw(ws, f"A{r}:E{r}", f"BALANCETE {mes_upper} {ano}",
        font=Font(bold=True, name="Calibri", size=11),
        fill=_YELLOW, align=_CCENT, border=_B_DOUBLE)
    r += 1

    # ── Header de despesas ──
    _mw(ws, f"A{r}:B{r}", "Natureza das Despesas",  font=_BOLD, border=_B_DOUBLE, align=_CCENT)
    _mw(ws, f"C{r}:E{r}", "Valor Realizado (R$)", font=_BOLD, border=_B_DOUBLE, align=_CCENT)
    r += 1

    _w(ws, r, 1, "Nº",        font=_BOLD, border=_B_THIN, align=_CCENT)
    _w(ws, r, 2, "Descrição", font=_BOLD, border=_B_THIN, align=_CCENT)
    _mw(ws, f"C{r}:E{r}", "Total", font=_BOLD, border=_B_THIN, align=_CCENT)
    r += 1

    # ── DESPESAS CORRENTES ──
    r_corr = r
    _mw(ws, f"A{r}:B{r}", "DESPESAS CORRENTES", font=_BOLD, border=_B_DOUBLE, align=_CLEFT)
    r += 1

    r_corr_ini = r
    for i, item in enumerate(dados["correntes"]):
        _linha_item(ws, r, i + 1, item["descricao"], item["valor"])
        r += 1
    r_corr_fim = r - 1
    _set_total_formula(ws, r_corr, r_corr_ini, r_corr_fim)

    # ── SERVIÇOS ──
    r_serv = r
    _mw(ws, f"A{r}:B{r}", "SERVIÇOS", font=_BOLD, border=_B_DOUBLE, align=_CLEFT)
    r += 1

    r_serv_ini = r
    for i, item in enumerate(dados["servicos"]):
        _linha_item(ws, r, i + 1, item["descricao"], item["valor"])
        r += 1
    r_serv_fim = r - 1
    _set_total_formula(ws, r_serv, r_serv_ini, r_serv_fim)

    r += 1  # spacer

    # ── TOTAL DESPESAS ──
    r_total_desp = r
    _mw(ws, f"A{r}:B{r}", "A. TOTAL DAS DESPESAS (a)", font=_BOLD, border=_B_DOUBLE, align=_CLEFT)
    _mw(ws, f"C{r}:E{r}", f"=C{r_corr}+C{r_serv}",
        font=_BOLD, border=_B_DOUBLE, align=_CRIGHT, num_fmt=_FMT_M)
    r += 2  # spacer

    # ── Header de receitas ──
    _mw(ws, f"A{r}:B{r}", "Natureza das Receitas",  font=_BOLD, border=_B_DOUBLE, align=_CCENT)
    _mw(ws, f"C{r}:E{r}", "Valor Realizado (R$)", font=_BOLD, border=_B_DOUBLE, align=_CCENT)
    r += 1

    _w(ws, r, 1, "Nº",        font=_BOLD, border=_B_THIN, align=_CCENT)
    _w(ws, r, 2, "Descrição", font=_BOLD, border=_B_THIN, align=_CCENT)
    _mw(ws, f"C{r}:E{r}", "Total", font=_BOLD, border=_B_THIN, align=_CCENT)
    r += 1

    # ── Itens de receita ──
    r_rec_ini = r
    for i, item in enumerate(dados["receitas"]):
        _linha_item(ws, r, i + 1, item["descricao"], item["valor"])
        r += 1
    r_rec_fim = r - 1
    r += 1  # spacer

    # ── TOTAL RECEITAS ──
    r_total_rec = r
    _mw(ws, f"A{r}:B{r}", "B. TOTAL DAS RECEITAS (b)", font=_BOLD, border=_B_DOUBLE, align=_CLEFT)
    _mw(ws, f"C{r}:E{r}", f"=SUM(C{r_rec_ini}:C{r_rec_fim})",
        font=_BOLD, border=_B_DOUBLE, align=_CRIGHT, num_fmt=_FMT_M)
    r += 2  # spacer

    # ── SALDO DO MÊS ──
    _mw(ws, f"A{r}:B{r}", f"SALDO DE {mes_upper} (RECEITAS - DESPESAS)",
        font=_BOLD, border=_B_DOUBLE, align=_CLEFT)
    _mw(ws, f"C{r}:E{r}", f"=C{r_total_rec}-C{r_total_desp}",
        font=_BOLD, border=_B_DOUBLE, align=_CRIGHT, num_fmt=_FMT_M)
    r += 2  # spacer

    # ── SALDO ANTERIOR ──
    _mw(ws, f"A{r}:B{r}", f"Saldo Anterior — {dados['data_ant_str']}",
        font=_BOLD, fill=_YELLOW, border=_B_DOUBLE, align=_CLEFT)
    _mw(ws, f"C{r}:E{r}", "", fill=_YELLOW, border=_B_DOUBLE)
    r += 1

    _w(ws, r, 1, "",               border=_B_THIN, align=_CCENT)
    _w(ws, r, 2, "Conta Bancária", font=_BOLD, border=_B_THIN, align=_CCENT)
    _mw(ws, f"C{r}:E{r}", "Valor", font=_BOLD, border=_B_THIN, align=_CCENT)
    r += 1

    r_sant_ini = r
    _linha_item(ws, r, "", "Saldo Bancário (C/C)",      dados["saldo_ant_cc"]);   r += 1
    _linha_item(ws, r, "", "Saldo Bancário (Poupança)", dados["saldo_ant_poup"]); r += 1
    _linha_item(ws, r, "", "Caixa",                     dados["caixa_ant"]);      r += 1
    r_sant_fim = r - 1

    r_per_ant = r
    _mw(ws, f"A{r}:B{r}", "C. SALDO DO PERÍODO ANTERIOR",
        font=_BOLD, border=_B_DOUBLE, align=_CLEFT)
    _mw(ws, f"C{r}:E{r}", f"=SUM(C{r_sant_ini}:C{r_sant_fim})",
        font=_BOLD, border=_B_DOUBLE, align=_CRIGHT, num_fmt=_FMT_M)
    r += 2  # spacer

    # ── SALDO FINAL ──
    _mw(ws, f"A{r}:E{r}", f"Saldo em {dados['data_fin_str']}",
        font=_BOLD, fill=_YELLOW, border=_B_DOUBLE, align=_CLEFT)
    r += 1

    _w(ws, r, 1, "",               border=_B_THIN, align=_CCENT)
    _w(ws, r, 2, "Conta Bancária", font=_BOLD, border=_B_THIN, align=_CCENT)
    _mw(ws, f"C{r}:E{r}", "Valor", font=_BOLD, border=_B_THIN, align=_CCENT)
    r += 1

    r_sfin_ini = r
    _linha_item(ws, r, "", "Saldo Bancário (C/C)",      dados["saldo_fin_cc"]);   r += 1
    _linha_item(ws, r, "", "Saldo Bancário (Poupança)", dados["saldo_fin_poup"]); r += 1
    _linha_item(ws, r, "", "Caixa",                     dados["caixa_fin"]);      r += 1
    r_sfin_fim = r - 1

    _mw(ws, f"A{r}:B{r}", "D. SALDO DISPONÍVEL",
        font=_BOLD, border=_B_DOUBLE, align=_CLEFT)
    _mw(ws, f"C{r}:E{r}", f"=SUM(C{r_sfin_ini}:C{r_sfin_fim})",
        font=_BOLD, border=_B_DOUBLE, align=_CRIGHT, num_fmt=_FMT_M)
    r += 1

    # ── Assinatura e conselho ──
    hoje = date.today()
    _mw(ws, f"A{r}:E{r}",
        f"Rio de Janeiro, {hoje.day} de {MESES_PT[hoje.month]} de {hoje.year}.",
        font=_NORM, align=_CCENT)
    r += 1

    _mw(ws, f"A{r}:E{r}",
        "____________________________________\nAdministrador do Condomínio",
        font=_NORM, align=_CCENT)
    r += 1

    _mw(ws, f"A{r}:E{r}", "Parecer do Conselho Consultivo:", font=_BOLD, align=_CLEFT)
    r += 1

    _mw(ws, f"A{r}:E{r}",
        "( ) Aprovado   ( ) Aprovado com ressalvas   ( ) Reprovado",
        font=_NORM, align=_CLEFT)
    r += 1

    _mw(ws, f"A{r}:B{r}",
        "____________________________________\nConselheiro",
        font=_NORM, align=_CCENT)
    _mw(ws, f"C{r}:E{r}",
        "____________________________________\nConselheiro",
        font=_NORM, align=_CCENT)
    r += 2

    # ── Rodapé ──
    _mw(ws, f"A{r}:E{r}",
        f"Relatório financeiro e administrativo — {mes_upper}/{ano}  |  Página 1 de 1",
        font=Font(italic=True, name="Calibri", size=9), align=_CCENT)

    wb.save(output_path)
    print(f"\n✅ Balancete salvo em: {output_path}")


# ═══════════════════════════════════════════════════════════════════
# FUNÇÃO PRINCIPAL
# ═══════════════════════════════════════════════════════════════════

def gerar_balancete(
    base_dir:              Path,
    mes_nome:              str,
    mes_num:               int,
    ano:                   int,
    caixa_fin:             float,
    caixa_ant:             float = None,
    arq_tacon_tuv:         Path  = None,
    inad_ant_dados:        dict  = None,
    inadimplente_forcado:  dict  = None,
) -> dict:
    """
    Etapa 3: gera o balancete mensal do Condomínio Barão de Mesquita.

    Parâmetros:
      base_dir             — pasta raiz com as subpastas Extratos/, etc.
      mes_nome             — nome do mês em português minúsculo (ex: "março")
      mes_num              — número do mês (ex: 3)
      ano                  — ano (ex: 2026)
      caixa_fin            — saldo em espécie no fim do mês
      caixa_ant            — saldo em espécie no início do mês (None = igual ao final)
      arq_tacon_tuv        — Path para o arquivo Cond TACON_TUV.xlsx (obrigatório)
      inad_ant_dados       — dict com inadimplentes do mês anterior (None = busca em base_dir)
      inadimplente_forcado — dict da unidade selecionada manualmente (resolve ambiguidade)

    Retorna:
      dict com arquivo_gerado, inadimplentes_arquivo, ambiguidade, candidatos e resumo.
    """
    base_dir = Path(base_dir)

    print("=" * 60)
    print(f"ETAPA 3 — GERAÇÃO DO BALANCETE {mes_nome.upper()} {ano}")
    print("=" * 60)

    # ── Arquivos de entrada ──
    arq_extrato_cc   = base_dir / "Extratos" / "Entrada" / "Extrato Corrente.html"
    arq_extrato_poup = base_dir / "Extratos" / "Entrada" / "Extrato Poupança.html"
    arq_despesas_ref = base_dir / "Extratos" / "Entrada" / "despesas_referencia.xlsx"

    for arq in [arq_extrato_cc, arq_extrato_poup, arq_despesas_ref]:
        if not arq.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {arq}")

    if arq_tacon_tuv is None:
        raise ValueError("arq_tacon_tuv é obrigatório. Informe o caminho da planilha TACON_TUV.")
    if not Path(arq_tacon_tuv).exists():
        raise FileNotFoundError(f"Planilha TACON_TUV não encontrada: {arq_tacon_tuv}")

    # ── Carregar config ──
    config = ler_config()
    print(f"\n✅ Config carregado: {len(config['contas'])} contas mapeadas")

    # ── Inadimplentes do mês anterior ──
    mes_ant_num, delta = MES_ANT[mes_num]
    ano_ant = ano + delta
    inad_ant = ler_inadimplentes_anteriores(base_dir, mes_ant_num, ano_ant, inad_ant_dados)
    ref_anterior_principal = sum(
        float(i.get("total", 0)) for i in inad_ant.get("inadimplentes", [])
    )
    print(f"\n✅ Inadimplência anterior: R$ {ref_anterior_principal:.2f} "
          f"({len(inad_ant['inadimplentes'])} unidade(s))")

    # ── Tabela TACON/TUV ──
    df_tacon, tacon_esp, tuv_esp = ler_tacon_tuv(Path(arq_tacon_tuv))
    print(f"\n✅ TACON esperado: R$ {tacon_esp:.2f} | TUV esperado: R$ {tuv_esp:.2f}")

    # ── Ler extratos ──
    print("\nLendo extrato corrente...")
    rows_cc = ler_extrato(arq_extrato_cc, "corrente")
    print(f"   {len(rows_cc)} linhas")

    print("Lendo extrato poupança...")
    rows_poup = ler_extrato(arq_extrato_poup, "poupanca")
    print(f"   {len(rows_poup)} linhas")

    # ── Saldos ──
    saldo_ant_cc,   saldo_fin_cc   = calcular_saldos_cc(rows_cc,   mes_num, ano)
    saldo_ant_poup, saldo_fin_poup = calcular_saldos_poupanca(rows_poup, mes_num, ano)

    if caixa_ant is None:
        caixa_ant = caixa_fin

    print(f"\n── Saldos ──")
    print(f"   CC anterior:       R$ {saldo_ant_cc:,.2f}")
    print(f"   CC final:          R$ {saldo_fin_cc:,.2f}")
    print(f"   Poupança anterior: R$ {saldo_ant_poup:,.2f}")
    print(f"   Poupança final:    R$ {saldo_fin_poup:,.2f}")

    # ── Poupança: juros e IRRF ──
    poup = calcular_poupanca(rows_poup, config, mes_num, ano)
    print(f"\n── Poupança ──")
    print(f"   Juros recebidos: R$ {poup['juros']:,.2f}")
    print(f"   IRRF pago:       R$ {poup['irrf']:,.2f}")

    # ── Despesas bancárias CC ──
    desp_banc_cc = calcular_despesas_bancarias_cc(rows_cc, config, mes_num, ano)
    print(f"\n── Despesas Bancárias ──")
    print(f"   CC:       R$ {desp_banc_cc:,.2f}")
    print(f"   Poupança: R$ {poup['irrf']:,.2f}")

    # ── COBs e receitas extras ──
    cob_data       = calcular_cobs_e_extras(rows_cc, config, mes_num, ano)
    cob_values     = cob_data["cob_values"]
    receitas_extras = cob_data["receitas_extras"]
    print(f"\n── COBs ──")
    print(f"   {len(cob_values)} pagamentos | Total: R$ {sum(cob_values):,.2f}")

    # ── TACON / TUV / Multa / Ref anterior ──
    print("\n── Calculando TACON/TUV ──")
    cob_info = calcular_tacon_tuv(
        cob_values             = cob_values,
        ref_anterior_principal = ref_anterior_principal,
        tacon_esperado         = tacon_esp,
        tuv_esperado           = tuv_esp,
        df_tacon_tuv           = df_tacon,
        inadimplente_forcado   = inadimplente_forcado,
    )
    print(f"   TACON:        R$ {cob_info['tacon']:,.2f}")
    print(f"   TUV:          R$ {cob_info['tuv']:,.2f}")
    print(f"   Multa:        R$ {cob_info['multa']:,.2f}")
    print(f"   Ref anterior: R$ {cob_info['ref_anterior_total']:,.2f}")

    # Se há ambiguidade, retornar agora para que a UI possa resolver
    if cob_info["ambiguidade"]:
        print("\n⚠️  Processamento pausado: inadimplência ambígua aguarda confirmação manual.")
        return {
            "arquivo_gerado":        None,
            "inadimplentes_arquivo": None,
            "ambiguidade":           True,
            "candidatos":            cob_info["candidatos"],
            "shortfall":             cob_info["shortfall"],
            "resumo":                None,
        }

    # ── Classificar despesas ──
    print("\nClassificando despesas de referência...")
    desp = ler_despesas_referencia(arq_despesas_ref, config)

    # Adicionar despesas bancárias às correntes
    ultimo_dia = calendar.monthrange(ano, mes_num)[1]
    desp["correntes"].append({
        "descricao": config["despesas_bancarias"]["cc_descricao"],
        "valor":     desp_banc_cc,
        "data":      date(ano, mes_num, ultimo_dia),
    })
    desp["correntes"].append({
        "descricao": config["despesas_bancarias"]["poupanca_descricao"],
        "valor":     poup["irrf"],
        "data":      date(ano, mes_num, ultimo_dia),
    })

    print(f"   Correntes: {len(desp['correntes'])} itens")
    print(f"   Serviços:  {len(desp['servicos'])} itens")

    # ── Montar receitas (ordem fixa) ──
    mes_ant_nome = MESES_PT[mes_ant_num]
    receitas = []
    receitas.append({"descricao": "Juros de Aplicação Financeira (Poupança)",
                     "valor": poup["juros"]})
    receitas.append({"descricao": "Taxa de Condomínio - TACON",
                     "valor": cob_info["tacon"]})
    receitas.append({"descricao": "Taxa de Utilização de Veículos - TUV",
                     "valor": cob_info["tuv"]})
    if cob_info["multa"] > 0.001:
        receitas.append({"descricao": "Multa - Pagamentos com atraso no mês corrente",
                         "valor": cob_info["multa"]})
    if cob_info["ref_anterior_total"] > 0.001:
        receitas.append({
            "descricao": (f"TACON + TUV + TAXA + MULTA + JUROS "
                          f"referentes a {mes_ant_nome.capitalize()}/{ano_ant}"),
            "valor": cob_info["ref_anterior_total"],
        })
    for extra in receitas_extras:
        descr = config["receitas_extras"].get("descricao_padrao", extra["hist"])
        receitas.append({"descricao": descr, "valor": extra["val"]})

    # ── Datas para o Excel ──
    data_fin_str = date(ano, mes_num, ultimo_dia).strftime("%d/%m/%Y")
    ultimo_dia_ant = calendar.monthrange(ano_ant, mes_ant_num)[1]
    data_ant_str   = date(ano_ant, mes_ant_num, ultimo_dia_ant).strftime("%d/%m/%Y")

    dados_excel = {
        "correntes":    desp["correntes"],
        "servicos":     desp["servicos"],
        "receitas":     receitas,
        "saldo_ant_cc":   saldo_ant_cc   or 0.0,
        "saldo_ant_poup": saldo_ant_poup or 0.0,
        "caixa_ant":      caixa_ant,
        "saldo_fin_cc":   saldo_fin_cc   or 0.0,
        "saldo_fin_poup": saldo_fin_poup or 0.0,
        "caixa_fin":      caixa_fin,
        "data_ant_str":   data_ant_str,
        "data_fin_str":   data_fin_str,
    }

    # ── Gerar XLSX ──
    nome_arquivo = f"balancete_{mes_nome[:3].lower()}{ano}.xlsx"
    output_path  = base_dir / nome_arquivo
    print(f"\nGerando Excel: {nome_arquivo}")
    gerar_excel_balancete(dados_excel, mes_nome, mes_num, ano, output_path)

    # ── Resumo ──
    total_desp_corr = sum(i["valor"] for i in desp["correntes"])
    total_serv      = sum(i["valor"] for i in desp["servicos"])
    total_desp      = round(total_desp_corr + total_serv, 2)
    total_rec       = round(sum(i["valor"] for i in receitas), 2)
    saldo_mes       = round(total_rec - total_desp, 2)
    saldo_ant_total = round((saldo_ant_cc or 0) + (saldo_ant_poup or 0) + caixa_ant, 2)
    saldo_fin_total = round((saldo_fin_cc or 0) + (saldo_fin_poup or 0) + caixa_fin, 2)

    print(f"\n{'─'*50}")
    print(f"  Despesas Correntes:  R$ {total_desp_corr:>12,.2f}")
    print(f"  Serviços:            R$ {total_serv:>12,.2f}")
    print(f"  A. Total Despesas:   R$ {total_desp:>12,.2f}")
    print(f"  B. Total Receitas:   R$ {total_rec:>12,.2f}")
    print(f"  Saldo do Mês (B-A):  R$ {saldo_mes:>12,.2f}")
    print(f"  C. Saldo Anterior:   R$ {saldo_ant_total:>12,.2f}")
    print(f"  D. Saldo Disponível: R$ {saldo_fin_total:>12,.2f}")
    print(f"{'─'*50}")

    # ── Salvar inadimplentes do mês atual ──
    inadimplentes_mes = []
    inad = cob_info.get("inadimplente_mes")
    if cob_info["shortfall"] > 0.01 and isinstance(inad, dict):
        col_u = col_t = col_tv = None
        for c in df_tacon.columns:
            cu = c.strip().upper()
            if cu == "UNIDADE": col_u  = c
            elif cu == "TACON": col_t  = c
            elif cu == "TUV":   col_tv = c
        inadimplentes_mes.append({
            "unidade":        str(inad.get(col_u, "?")) if col_u else "?",
            "tacon":          float(inad.get(col_t,  0) or 0),
            "tuv":            float(inad.get(col_tv, 0) or 0),
            "total":          cob_info["shortfall"],
            "identificacao":  "manual" if inadimplente_forcado else "automatica",
        })

    arq_inad_saida = base_dir / f"inadimplentes_{mes_nome}{ano}.json"
    with open(arq_inad_saida, "w", encoding="utf-8") as f:
        json.dump({
            "competencia":   f"{ano}-{mes_num:02d}",
            "mes_nome":      mes_nome,
            "inadimplentes": inadimplentes_mes,
        }, f, ensure_ascii=False, indent=2)
    print(f"\n✅ Inadimplentes do mês salvos em: {arq_inad_saida.name}")

    return {
        "arquivo_gerado":        output_path,
        "inadimplentes_arquivo": arq_inad_saida,
        "ambiguidade":           False,
        "candidatos":            [],
        "shortfall":             cob_info["shortfall"],
        "resumo": {
            "total_despesas_correntes": total_desp_corr,
            "total_servicos":           total_serv,
            "total_despesas":           total_desp,
            "total_receitas":           total_rec,
            "saldo_mes":                saldo_mes,
            "saldo_anterior":           saldo_ant_total,
            "saldo_disponivel":         saldo_fin_total,
            "tacon":                    cob_info["tacon"],
            "tuv":                      cob_info["tuv"],
            "multa":                    cob_info["multa"],
            "ref_anterior":             cob_info["ref_anterior_total"],
            "juros_poupanca":           poup["juros"],
            "irrf_poupanca":            poup["irrf"],
            "despesas_bancarias_cc":    desp_banc_cc,
        },
    }


# ═══════════════════════════════════════════════════════════════════
# EXECUÇÃO DIRETA (para testes via CLI)
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import sys

    BASE = Path("C:/Projetos pessoais/Automação condominio")
    TACON = BASE / "Melhoria balancete" / "Cond TACON_TUV.xlsx"

    resultado = gerar_balancete(
        base_dir      = BASE,
        mes_nome      = "março",
        mes_num       = 3,
        ano           = 2026,
        caixa_fin     = -50.15,
        caixa_ant     = -50.15,
        arq_tacon_tuv = TACON,
    )

    print(f"\n{'═'*60}")
    if resultado["arquivo_gerado"]:
        print("ARQUIVO GERADO:", resultado["arquivo_gerado"])
    if resultado["ambiguidade"]:
        print("⚠️  ATENÇÃO: Inadimplência ambígua — confirmação manual necessária.")
        for c in resultado["candidatos"]:
            print(f"   Candidato: {c}")
    print("Concluído.")
